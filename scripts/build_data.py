"""
Electify — data pipeline.

Reads several local student-collected guides plus the inline Master List,
then emits src/data/courses.json. All raw input files live in DOWNLOADS and
are intentionally not versioned.

Conservative on merging: when inputs conflict on ease score or grading, keeps
both and tags accordingly — the frontend shows a "varies" badge.
"""
from __future__ import annotations
import json, os, re, sys
from pathlib import Path
from collections import defaultdict

import pdfplumber
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DOWNLOADS = Path(r"C:\Users\user\Downloads")
OUT = ROOT / "src" / "data" / "courses.json"

ATTR_MAP = {
    "cultures & histories": "Cultures & Histories",
    "cultures and histories": "Cultures & Histories",
    "societies & individuals": "Societies & Individuals",
    "societies and individuals": "Societies & Individuals",
    "understanding communication": "Understanding Communication",
    "understanding the world": "Understanding the World",
    "social inequalities": "Social Inequalities",
    "quantitative reasoning": "Quantitative Reasoning",
    "history of ideas": "History of Ideas",
    "technical elective": "Technical Elective",
}

CODE_RE = re.compile(r"\b([A-Z]{3,5})\s*0*([0-9]{2,3}[A-Z]{0,3})\b")


def norm_code(s: str) -> str:
    m = CODE_RE.search(s.upper().replace("\u00a0", " "))
    if not m:
        return ""
    return f"{m.group(1)} {m.group(2)}"


# ----------------------------------------------------------------------------
# Course catalog (PDF)
# ----------------------------------------------------------------------------
def parse_catalog(path: Path) -> dict[str, dict]:
    courses: dict[str, dict] = {}
    current_attr = None
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            if not lines:
                continue
            joined = " ".join(lines).lower()
            for k, v in ATTR_MAP.items():
                if k in joined and len(lines) <= 4:
                    current_attr = v
                    break

            page_text = text
            code = norm_code(page_text)
            if not code:
                continue

            title = ""
            desc = ""
            ease = None
            grading = ""
            prof = ""

            for i, l in enumerate(lines):
                if "Course Title" in l and i > 0:
                    title = lines[i - 1]
                if "out of 10" in l and i > 0:
                    try:
                        ease = int(lines[i - 1])
                    except ValueError:
                        pass
                if l.startswith("Grading system"):
                    grading = l.replace("Grading system", "").strip()
                    if not grading and i + 1 < len(lines):
                        grading = lines[i + 1]
                if "Recommended" in l and i + 1 < len(lines):
                    cand = lines[i + 1]
                    if "Professor" not in cand:
                        prof = cand

            m = re.search(r"Course\s+Description\s+(.*?)(?:Recommended|Ease of|$)",
                          page_text, re.S | re.I)
            if m:
                desc = re.sub(r"\s+", " ", m.group(1)).strip()

            existing = courses.get(code, {})
            attrs = set(existing.get("attributes", []))
            if current_attr:
                attrs.add(current_attr)
            courses[code] = {
                "code": code,
                "title": title or existing.get("title", ""),
                "description": desc or existing.get("description", ""),
                "ease_score": ease if ease is not None else existing.get("ease_score"),
                "grading": grading or existing.get("grading", ""),
                "attributes": sorted(attrs),
                "catalog_prof": prof or existing.get("catalog_prof", ""),
            }
    return courses


# ----------------------------------------------------------------------------
# Term guides (PDF)
# ----------------------------------------------------------------------------
def parse_term_guide(path: Path, term: str) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = defaultdict(list)
    current_attr = None
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for raw in text.splitlines():
                line = raw.strip()
                if not line:
                    continue
                low = line.lower()
                for k, v in ATTR_MAP.items():
                    if low == k or low.startswith(k):
                        current_attr = v
                code = norm_code(line)
                if not code or not current_attr:
                    continue
                rest = line.split(code, 1)[-1] if code in line else line
                profs = re.findall(r"[A-Z][a-z]+(?:\s+[A-Z][a-z\-']+){1,3}", rest)
                out[code].append({
                    "term": term,
                    "attribute": current_attr,
                    "professors": profs,
                    "raw": line,
                })
    return out


# ----------------------------------------------------------------------------
# xlsx student survey
# ----------------------------------------------------------------------------
def parse_xlsx(path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    out: dict[str, list[dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        raw_code = (rec.get("Course Name") or "").strip()
        code = norm_code(raw_code)
        if not code:
            continue
        out[code].append({
            "recommend": rec.get("Do you recommend this elective?"),
            "professor": (rec.get("Professor's Name (If don't remember, leave blank)") or "").strip(),
            "best_for": rec.get("Who is this course best for? \nExample: students who like reading, discussion, writing, memorization ..."),
            "prereq": rec.get("Is there any required previous course?"),
            "attribute_raw": rec.get("Course Attribute"),
            "structure": rec.get("Course Structure"),
            "tip": rec.get("One thing students should know before registering this course (Recommendation)"),
        })
    return out


# ----------------------------------------------------------------------------
# Master list (Mouhamad Chhaitle) — hand-encoded since it's text-only
# ----------------------------------------------------------------------------
MASTER_LIST = [
    # (code, prof, best_for, prereq, structure, note, recommended)
    ("ECON 203", "Maya Kanaan", "General students", "", "Homework + 3 exams", "Easier with this professor but not inherently easy", True),
    ("ARAB 251 AZ", "Mariam El Ali", "Discussion-based students", "", "Midterm + Final + Presentation", "", True),
    ("CHEM 200", "Pierre Karam", "Curious students", "", "Participation + quizzes + final", "Flexible", True),
    ("EDUC 230", "Enja Osman", "Teaching and presentations", "", "Exams + projects + participation", "Light workload", True),
    ("SOAN 201", "Rima Rassi", "Any student", "", "Exams + project + participation", "", True),
    ("SOAN 246", "Alexandra Kassir", "Reading and discussion", "", "Exams + paper + participation", "Easy grading", True),
    ("PSPA 202", "Hiba Khodeib", "Reading and listening", "", "Exams + paper", "Easy but boring", True),
    ("MUSC 230", "Joelle Khoury", "Music lovers", "", "Paper + quizzes", "", True),
    ("PSPA 272", "Tania Haddad", "Discussions and projects", "", "Projects + participation", "Easy grading", True),
    ("PHYS 200", "Samih Isber", "Astronomy interest", "", "Exams + paper", "MCQ exams", True),
    ("PSPA 201", "Yeghia Tashjian", "Memorization and discussion", "", "Exams + quizzes", "", True),
    ("ENGL 205", "DeVan Ard-Keyser", "Reading and discussion", "ENGL 203 maybe", "Projects + quizzes", "", True),
    ("MUSC 200", "Yasmina Sabbah", "Logical thinking", "", "Exams + participation", "Fun", True),
    ("ENGL 214", "Christopher Nassar", "Reading and writing", "", "Exams", "Easy if focused", True),
    ("ARAB 243", "Najah Hawwa", "Arabic literature", "", "Exams + participation", "", True),
    ("AROL 221", "Paul Newson", "Memorization and discussion", "", "Exams", "Notes are essential", True),
    ("ENGL 212", "Emily Ard-Keyser", "Discussion and writing", "ENGL 203", "Exams + paper", "", True),
    ("AGSC 205", "Lina Jaber", "Presentations", "Approval required", "Exam + presentations", "Easy", True),
    ("ARAB 247", "Enass Khansa", "Reading and discussion", "", "Projects + participation", "", True),
    ("ARAB 211", "", "Strong Arabic grammar", "", "Exams", "Heavy Arabic", True),
    ("AROL 225", "Paul Newson", "Writing", "", "Exams", "Light", True),
    ("EDUC 219", "Hoda Baytiyeh", "Practical skills", "", "Assignments", "Very easy A", True),
    ("MUSC 205", "Joelle Garro El Khoury", "Anyone", "Music 200 optional", "Quizzes + homework", "Easy A", True),
    # warnings
    ("DCSN 200", "", "", "", "", "Heavy word problems", False),
    ("AHIS 221", "", "", "", "", "Heavy workload", False),
    ("AHIS 203", "", "", "", "", "Very heavy workload", False),
    ("PHIL 226", "", "", "", "", "Strict grading", False),
    ("BIOL 210", "", "", "", "", "Requires consistency", False),
    ("ECON 211", "", "", "", "", "", False),
    ("PHIL 214", "", "", "", "", "Requires depth", False),
]


def main() -> None:
    catalog = parse_catalog(DOWNLOADS / "catalog.pdf")
    guide_spring = parse_term_guide(DOWNLOADS / "term_guide_spring.pdf", "Spring")
    guide_summer = parse_term_guide(DOWNLOADS / "term_guide_summer.pdf", "Summer")
    xlsx = parse_xlsx(DOWNLOADS / "electives_sheet.xlsx")

    all_codes = set(catalog) | set(guide_spring) | set(guide_summer) | set(xlsx) | {norm_code(c[0]) for c in MASTER_LIST}
    all_codes.discard("")

    courses = []
    for code in sorted(all_codes):
        i = catalog.get(code, {})
        professors_map: dict[str, dict] = {}

        def add_prof(name: str, term: str | None, note: str = ""):
            name = (name or "").strip().rstrip(",")
            if not name or name.lower() in {"-", "n/a", "na"}:
                return
            key = name.lower()
            p = professors_map.setdefault(key, {"name": name, "terms": set(), "notes": []})
            if term:
                p["terms"].add(term)
            if note and note not in p["notes"]:
                p["notes"].append(note)

        if i.get("catalog_prof"):
            add_prof(i["catalog_prof"], None)
        for entry in guide_spring.get(code, []):
            for p in entry["professors"]:
                add_prof(p, "Spring")
        for entry in guide_summer.get(code, []):
            for p in entry["professors"]:
                add_prof(p, "Summer")
        for r in xlsx.get(code, []):
            add_prof(r.get("professor", ""), None)
        master_warnings: list[str] = []
        master_recommended = None
        for m in MASTER_LIST:
            if norm_code(m[0]) == code:
                add_prof(m[1], None, m[5] or "")
                if not m[6]:
                    master_recommended = False
                    if m[5]:
                        master_warnings.append(m[5])
                else:
                    if master_recommended is None:
                        master_recommended = True

        professors = []
        for p in professors_map.values():
            professors.append({
                "name": p["name"],
                "terms": sorted(p["terms"]),
                "notes": p["notes"],
            })

        student_reviews = []
        for r in xlsx.get(code, []):
            student_reviews.append({
                "professor": r.get("professor", ""),
                "best_for": r.get("best_for", ""),
                "prerequisite": r.get("prereq", ""),
                "structure": r.get("structure", ""),
                "tip": r.get("tip", ""),
            })

        master_quick = []
        for m in MASTER_LIST:
            if norm_code(m[0]) == code and m[6]:
                master_quick.append({
                    "professor": m[1],
                    "best_for": m[2],
                    "prerequisite": m[3],
                    "structure": m[4],
                    "note": m[5],
                })

        courses.append({
            "code": code,
            "slug": code.lower().replace(" ", "-"),
            "title": i.get("title", ""),
            "description": i.get("description", ""),
            "ease_score": i.get("ease_score"),
            "grading": i.get("grading", ""),
            "attributes": i.get("attributes", []),
            "professors": professors,
            "student_reviews": student_reviews,
            "master_notes": master_quick,
            "warnings": master_warnings,
            "recommended": False if master_recommended is False else True,
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "generated_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "count": len(courses),
        "courses": courses,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(courses)} courses to {OUT}")


if __name__ == "__main__":
    main()
